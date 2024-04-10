import datetime
import os
import sys

from openpyxl import Workbook
from tqdm import tqdm

# openpyxl.worksheet.worksheet.Worksheet
from custom_logging import logger as clog
from parse_sql import parse_sql
from sql_connection import connection as scnxn
from stored_procedures import stored_procedures as sps

__version__ = "1.0.0"


def format_excel_file(excel_filepath: str):
    try:
        wb: Workbook = openpyxl.load_workbook(excel_filepath)
        ws = wb.active
        ws["A1"].font = openpyxl.styles.Font(bold=True, size=16)
        ws["B1"].font = openpyxl.styles.Font(bold=True, size=16)

        border = openpyxl.styles.Border(
            left=openpyxl.styles.Side(style="thin"),
            right=openpyxl.styles.Side(style="thin"),
            top=openpyxl.styles.Side(style="thin"),
            bottom=openpyxl.styles.Side(style="thin"),
        )

        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = openpyxl.styles.Alignment(wrap_text=True)
                cell.border = border

        ws.column_dimensions["A"].width = 75
        ws.column_dimensions["B"].width = 50

        wb.save(excel_filepath)

        return None

    except Exception as err:
        return err


def write_to_excel(
    output_filepath: str, entries: list[dict[str, set[str]]]
) -> Exception | None:
    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        headers = ["SP Name", "Table Names"]
        ws.append(headers)
        for entry in entries:
            sp_name, table_names = list(entry.items())[0]
            ws.append([sp_name, "\n".join(table_names)])
        wb.save(output_filepath)
        return None
    except Exception as err:
        return err


def main() -> None:
    print()
    clog.log(
        log_level=1, message=f"SQL Temp Table v{__version__} has started executing."
    )
    print()
    excel_entries: list[dict[str, set[str]]] = []
    timestamp = datetime.datetime.now().strftime("%Y_%m_%d_%H_%M")
    excel_filename = timestamp + "_output.xlsx"
    output_filepath = os.path.join(os.path.curdir, excel_filename)
    clog.log(log_level=1, message="Input SQL credentials:")
    server: str = input("Enter server address: ")
    database: str = input("Enter database name: ")
    username: str = input("Enter username: ")
    password: str = input("Enter password: ")

    err, connection = scnxn.get_connection(server, database, username, password)
    if err is not None:
        clog.log(3, f"Could not connect to SQL Database: {err}")
        clog.log(1, "Exiting program...")
        sys.exit(1)
    else:
        clog.log(log_level=1, message="Connection to the database succeeded.")

    if connection is not None:
        err, procs = sps.fetch_sp_names(connection=connection)
        if err is not None:
            clog.log(3, f"Could not fetch stored procedures from database: {err}")
        else:
            clog.log(log_level=1, message="Fetching stored procedures succeeded.")
            if procs is not None:
                clog.log(
                    log_level=1, message=f"Fetched `{len(procs)}` stored procedures"
                )
            else:
                clog.log(
                    log_level=2,
                    message="The list of stored procedure names was empty. Exiting...",
                )
                sys.exit(1)
    else:
        clog.log(log_level=2, message="Connection was unexpectedly empty. Exiting...")
        sys.exit(1)

    if procs is None:
        clog.log(
            log_level=2,
            message="The list of stored procedure names was empty. Exiting...",
        )
        sys.exit(1)

    pbar = tqdm(iterable=procs, colour="blue", dynamic_ncols=True, unit="SP")

    # for sp in procs:
    for sp in pbar:
        pbar.set_description("Processing `%s`" % sp)
        # clog.log(1, f"Processing SP: `{sp}`")
        err, sp_content = sps.fetch_sp_content(connection=connection, sp_name=sp)
        if err is not None:
            clog.log(
                log_level=3, message=f"Could not get contents of Stored Procedure: {sp}"
            )
            clog.log(log_level=1, message="Fetching content for next SP")
            continue
        else:
            if sp_content is None:
                clog.log(2, f"`{sp}` was empty")
                clog.log(log_level=1, message="Fetching content for next SP")
                continue
            err, fmt_sql = parse_sql.format_sql(sql_content=sp_content)
            if err is not None:
                clog.log(log_level=3, message="Could not format SQL")
                clog.log(log_level=1, message="Fetching content for next SP")
                continue
            else:
                if fmt_sql is None:
                    clog.log(
                        2,
                        "Formatted SQL was found to be empty. Moving on to the next SP.",
                    )
                    continue
                # clog.log(1, "Formatted SQL Contents:")
                # print(fmt_sql)
                undropped_temp_tables: set[str] = parse_sql.get_undropped_tables(
                    formatted_sql=fmt_sql
                )
                if len(undropped_temp_tables) > 0:
                    # clog.log(1, f"Temp Tables were found in SP: {sp}")
                    # clog.log(
                    #     1, f"Number of Undropped Temp Tables: {len(undropped_temp_tables)}")
                    # clog.log(
                    #     1, f"Undropped Temp Tables: {undropped_temp_tables}")
                    excel_entries.append({f"{sp}": undropped_temp_tables})
                    # sys.exit(0)
                else:
                    # clog.log(1, f"No TEMP Tables found in SP: {sp}")
                    continue
    err = write_to_excel(output_filepath=output_filepath, entries=excel_entries)
    if err is not None:
        clog.log(3, f"Could not write to excel {err}")
    else:
        clog.log(
            1, f"Entries written to excel file: {os.path.abspath(output_filepath)}"
        )

    err = format_excel_file(excel_filepath=output_filepath)
    if err is not None:
        clog.log(3, f"Error while formatting excel file: {err}")
    else:
        clog.log(1, "Excel file formatted successfully!")

    os.system(f"start explorer {os.path.dirname(os.path.abspath(output_filepath))}")


if __name__ == "__main__":
    main()
